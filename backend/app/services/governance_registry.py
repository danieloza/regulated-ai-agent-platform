from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook


REGISTRY_SHEETS = {
    "Policies": {
        "category": "policies",
        "headers": ["external_id", "version", "mode", "status", "owner", "description"],
        "required": ["external_id", "version", "mode", "status", "owner"],
        "enums": {"mode": {"standard", "strict", "regulated"}, "status": {"draft", "active", "retired"}},
    },
    "Risk Rules": {
        "category": "risk_rules",
        "headers": ["external_id", "factor", "weight", "threshold_band", "rationale", "owner"],
        "required": ["external_id", "factor", "weight", "threshold_band", "rationale"],
        "enums": {"threshold_band": {"low", "medium", "high"}},
    },
    "Security Evals": {
        "category": "security_evals",
        "headers": ["external_id", "input", "expected_decision", "category", "owner"],
        "required": ["external_id", "input", "expected_decision", "category"],
        "enums": {"expected_decision": {"allowed", "denied", "approval_required"}},
    },
    "Approved Sources": {
        "category": "approved_sources",
        "headers": ["external_id", "title", "classification", "owner", "review_date"],
        "required": ["external_id", "title", "classification", "owner"],
        "enums": {"classification": {"public", "internal", "confidential", "restricted"}},
    },
    "Control Owners": {
        "category": "control_owners",
        "headers": ["external_id", "team", "owner", "email", "review_date"],
        "required": ["external_id", "team", "owner", "email"],
        "enums": {},
    },
}

EXTERNAL_ID_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._-]{1,79}$")
EMAIL_RE = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")


def parse_registry_workbook(raw: bytes) -> dict:
    try:
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        return {"rows": [], "errors": [{"sheet": None, "row": None, "field": "file", "message": f"Workbook could not be read: {exc.__class__.__name__}."}]}

    rows: list[dict] = []
    errors: list[dict] = []
    for sheet_name, settings in REGISTRY_SHEETS.items():
        if sheet_name not in workbook.sheetnames:
            errors.append({"sheet": sheet_name, "row": None, "field": "sheet", "message": "Required worksheet is missing."})
            continue
        sheet = workbook[sheet_name]
        headers = [_normalize_header(sheet.cell(row=4, column=index + 1).value) for index in range(len(settings["headers"]))]
        if headers != settings["headers"]:
            errors.append(
                {
                    "sheet": sheet_name,
                    "row": 4,
                    "field": "headers",
                    "message": f"Expected headers: {', '.join(settings['headers'])}.",
                }
            )
            continue

        seen_ids: set[str] = set()
        for row_number, raw_values in enumerate(
            sheet.iter_rows(min_row=5, max_col=len(headers), values_only=True),
            start=5,
        ):
            values = [_normalize_value(value) for value in raw_values]
            if all(value in (None, "") for value in values):
                continue
            data = dict(zip(headers, values))
            row_errors = _validate_row(sheet_name, row_number, data, settings, seen_ids)
            errors.extend(row_errors)
            if not row_errors:
                external_id = str(data["external_id"])
                seen_ids.add(external_id)
                rows.append(
                    {
                        "sheet": sheet_name,
                        "category": settings["category"],
                        "row": row_number,
                        "external_id": external_id,
                        "data": data,
                    }
                )
            if len(rows) + len(errors) > 500:
                errors.append({"sheet": sheet_name, "row": row_number, "field": "file", "message": "Workbook exceeds the 500-row import limit."})
                return {"rows": rows, "errors": errors}
    return {"rows": rows, "errors": errors}


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower()


def _normalize_value(value: object) -> object:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def _validate_row(sheet: str, row: int, data: dict, settings: dict, seen_ids: set[str]) -> list[dict]:
    errors = []

    def add(field: str, message: str) -> None:
        errors.append({"sheet": sheet, "row": row, "field": field, "message": message})

    for field in settings["required"]:
        if data.get(field) in (None, ""):
            add(field, "Required value is missing.")

    external_id = str(data.get("external_id") or "")
    if external_id and not EXTERNAL_ID_RE.fullmatch(external_id):
        add("external_id", "Use 2-80 letters, numbers, dots, underscores, or hyphens.")
    if external_id in seen_ids:
        add("external_id", "Duplicate external_id in this worksheet.")

    for field, accepted in settings["enums"].items():
        value = data.get(field)
        if value not in (None, "") and str(value) not in accepted:
            add(field, f"Expected one of: {', '.join(sorted(accepted))}.")

    if sheet == "Risk Rules" and data.get("weight") not in (None, ""):
        weight = data["weight"]
        if isinstance(weight, bool) or not isinstance(weight, (int, float)) or int(weight) != weight or not 0 <= int(weight) <= 100:
            add("weight", "Weight must be an integer from 0 to 100.")
        else:
            data["weight"] = int(weight)

    if sheet == "Control Owners" and data.get("email") and not EMAIL_RE.fullmatch(str(data["email"])):
        add("email", "Enter a valid email address.")

    for field in ("review_date",):
        if data.get(field) and not _is_iso_date(str(data[field])):
            add(field, "Use an Excel date value or ISO date yyyy-mm-dd.")
    return errors


def _is_iso_date(value: str) -> bool:
    try:
        date.fromisoformat(value)
        return True
    except ValueError:
        return False
